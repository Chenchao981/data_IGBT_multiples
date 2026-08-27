"""Fail-closed workbook signatures for TMS FT DC adapters."""

from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import load_workbook


def _row_values(rows: list[tuple[object, ...]], row_index: int) -> set[str]:
    if row_index >= len(rows):
        return set()
    return {
        str(value).strip()
        for value in rows[row_index]
        if value is not None and str(value).strip()
    }


def validate_dc_workbook(
    file_path: str | Path,
    *,
    factory_name: str,
    unit_row_index: int,
    test_no_row_index: int,
    time_row_index: int | None,
) -> None:
    """Validate the exact header rows without loading measurement rows."""
    path = Path(file_path)
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"{factory_name} FT DC 只支持 XLSX: {path.name}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if workbook.sheetnames != ["Test Data"]:
            raise ValueError(
                f"{factory_name} FT DC 工作表必须且只能是 Test Data: {path.name}"
            )
        worksheet = workbook["Test Data"]
        max_row = max(test_no_row_index, unit_row_index, time_row_index or 0) + 1
        rows = list(
            worksheet.iter_rows(min_row=1, max_row=max_row, values_only=True)
        )
    finally:
        workbook.close()

    item_row = _row_values(rows, 1)
    if "Item" not in item_row or "CONT" not in item_row:
        raise ValueError(f"{factory_name} FT DC 第2行缺少 Item/CONT: {path.name}")
    if "Unit" not in _row_values(rows, unit_row_index):
        raise ValueError(
            f"{factory_name} FT DC Unit 行应为第{unit_row_index + 1}行: {path.name}"
        )
    test_row = _row_values(rows, test_no_row_index)
    if not ({"Test No.", "Test No"} & test_row):
        raise ValueError(
            f"{factory_name} FT DC Test No. 行应为第{test_no_row_index + 1}行: {path.name}"
        )
    if time_row_index is not None and "Time" not in _row_values(rows, time_row_index):
        raise ValueError(
            f"{factory_name} FT DC Time 行应为第{time_row_index + 1}行: {path.name}"
        )


def rewrite_manifest_factory(manifest_path: Path | None, *, code: str, name: str) -> None:
    """Correct the portable bundle identity produced by the mature base Cleaner."""
    if manifest_path is None or not manifest_path.is_file():
        raise ValueError(f"{name} FT DC 未生成散点图清单")
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    payload["factory"] = name
    payload["factory_code"] = code
    manifest_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def validate_complete_source_coverage(
    manifest_path: Path | None,
    *,
    expected_sources: set[str],
    factory_name: str,
) -> None:
    """Require every registered XLSX to contribute both data and source-specific Spec."""
    if manifest_path is None or not manifest_path.is_file():
        raise ValueError(f"{factory_name} FT DC 未生成散点图清单")
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest_sources = {
        str(value).strip()
        for value in payload.get("sources", [])
        if str(value).strip()
    }
    if manifest_sources != expected_sources:
        missing = sorted(expected_sources - manifest_sources)
        extra = sorted(manifest_sources - expected_sources)
        raise ValueError(
            f"{factory_name} FT DC 未完整处理所有登记文件: "
            f"缺少数据来源={missing}, 未登记来源={extra}"
        )

    spec_path = manifest_path.parent / "ft_scatter_spec.csv"
    if not spec_path.is_file():
        raise ValueError(f"{factory_name} FT DC 未生成逐来源Spec")
    with spec_path.open("r", encoding="utf-8-sig", newline="") as stream:
        reader = csv.DictReader(stream)
        if "Source_ID" not in (reader.fieldnames or []):
            raise ValueError(f"{factory_name} FT DC Spec缺少Source_ID")
        spec_sources = {
            str(row.get("Source_ID") or "").strip()
            for row in reader
            if str(row.get("Source_ID") or "").strip()
        }
    if spec_sources != expected_sources:
        missing = sorted(expected_sources - spec_sources)
        extra = sorted(spec_sources - expected_sources)
        raise ValueError(
            f"{factory_name} FT DC 未完整生成逐来源Spec: "
            f"缺少来源={missing}, 未登记来源={extra}"
        )
