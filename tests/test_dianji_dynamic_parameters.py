import gc
import tempfile
import unittest
import warnings
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from factories.dianji.dc_cleaner import DianjiDCCleaner
from factories.dianji.models import DianjiFormatError
from factories.dianji.pat_cleaner import build_raw_pat
from factories.dianji.powertech_parser import parse_powertech_file
from factories.dianji.powertech_xlsx_parser import parse_powertech_xlsx_file
from tests.test_dianji_powertech_cleaner import _make_source as _make_text_source
from tests.test_dianji_powertech_xlsx_cleaner import _make_source as _make_xlsx_source


def _append_text_items(path: Path, items: list[dict[str, object]]) -> Path:
    """Append synthetic PowerTECH Items without filling early-stop short rows."""
    lines = path.read_text(encoding="gb18030").splitlines()
    rows = [line.split("\t") for line in lines]
    item_row = next(row for row in rows if row and row[0] == "Item Name")
    old_width = len(item_row)
    existing_numbers = [
        int(value.split(" ", 1)[0])
        for value in item_row[3:]
        if value.strip()
    ]
    next_number = max(existing_numbers) + 1
    shaped_rows = {
        row[0]: row
        for row in rows
        if row
        and row[0]
        in {"Item Name", "Bias1", "Bias2", "Bias3", "Min Limit", "Max Limit", "Serial#"}
    }

    for offset, item in enumerate(items):
        number = next_number + offset
        shaped_rows["Item Name"].append(f"{number} {item['name']}")
        shaped_rows["Bias1"].append(str(item.get("bias1", "")))
        shaped_rows["Bias2"].append(str(item.get("bias2", "")))
        shaped_rows["Bias3"].append(str(item.get("bias3", "")))
        shaped_rows["Min Limit"].append(str(item.get("low", "")))
        shaped_rows["Max Limit"].append(str(item.get("high", "")))
        shaped_rows["Serial#"].append(str(item.get("unit", "")))

    for row in rows:
        if not row or not row[0].isdigit() or len(row) < old_width:
            continue
        serial = row[0]
        for item in items:
            values = item.get("values", {})
            assert isinstance(values, dict)
            row.append(str(values.get(serial, "")))

    path.write_text(
        "\r\n".join("\t".join(row) for row in rows),
        encoding="gb18030",
    )
    return path


def _append_xlsx_items(path: Path, items: list[dict[str, object]]) -> Path:
    workbook = load_workbook(path)
    sheet = workbook["Datalog"]
    existing_numbers = []
    for column in range(4, sheet.max_column + 1):
        value = sheet.cell(8, column).value
        if value:
            existing_numbers.append(int(str(value).split(" ", 1)[0]))
    next_number = max(existing_numbers) + 1

    for offset, item in enumerate(items):
        number = next_number + offset
        column = number + 3
        sheet.cell(8, column, f"{number} {item['name']}")
        sheet.cell(9, column, str(item.get("bias1", "")))
        sheet.cell(10, column, str(item.get("bias2", "")))
        sheet.cell(11, column, str(item.get("bias3", "")))
        sheet.cell(12, column, str(item.get("low", "")))
        sheet.cell(13, column, str(item.get("high", "")))
        sheet.cell(18, column, str(item.get("unit", "")))
        values = item.get("values", {})
        assert isinstance(values, dict)
        for row in range(19, sheet.max_row + 1):
            serial = sheet.cell(row, 1).value
            if str(serial) in values:
                sheet.cell(row, column, values[str(serial)])

    workbook.save(path)
    workbook.close()
    return path


def _matching_parameter(columns, token: str) -> str:
    matches = [str(column) for column in columns if token.casefold() in str(column).casefold()]
    if len(matches) != 1:
        raise AssertionError(f"expected one parameter containing {token!r}, got {matches}")
    return matches[0]


def _collect_xlsx_handles() -> None:
    # The Calamine/openpyxl combination can defer closing a read handle until
    # collection on Windows.  Collect before TemporaryDirectory cleanup while
    # keeping the regression output focused on parser behavior.
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", ResourceWarning)
        gc.collect()


class PowerTechTextDynamicParameterTests(unittest.TestCase):
    def _long_source(self, directory: Path, **identity) -> Path:
        source = _make_text_source(directory, **identity)
        return _append_text_items(
            source,
            [
                {
                    "name": "IDSS",
                    "bias1": "VDS=80.00V",
                    "unit": "uA",
                    "low": "0.100uA",
                    "high": "0.400uA",
                    "values": {"1": "0.125", "4": "0.250"},
                },
                {
                    "name": "DYN_LEAK",
                    "bias1": "VDS=70.00V",
                    "unit": "uA",
                    "low": "0.010uA",
                    "high": "0.500uA",
                    "values": {"1": "0.050", "4": "0.075"},
                },
                {
                    "name": "CONT_NEW",
                    "unit": "V",
                    "values": {"1": "1", "4": "1"},
                },
                {
                    "name": "CONTACT",
                    "unit": "V",
                    "values": {"1": "1", "4": "1"},
                },
                {
                    "name": "SAME",
                    "bias1": "M#=36",
                    "values": {"1": "1", "4": "1"},
                },
                {
                    "name": "DELAY",
                    "bias1": "Time=10.00ms",
                    "unit": "ms",
                    "values": {"1": "10", "4": "10"},
                },
            ],
        )

    def test_includes_known_and_unknown_right_side_items_and_excludes_controls(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as temp:
            parsed = parse_powertech_file(self._long_source(Path(temp)))

        self.assertIn("IDSS80(nA)", parsed.data.columns)
        self.assertEqual(
            parsed.data["IDSS80(nA)"].dropna().tolist(),
            [125.0, 250.0],
        )
        idss_spec = parsed.specs.loc[
            parsed.specs["Parameter"] == "IDSS80(nA)"
        ].iloc[0]
        self.assertEqual(idss_spec["Unit"], "nA")
        self.assertEqual(float(idss_spec["Low_Limit"]), 100.0)
        self.assertEqual(float(idss_spec["High_Limit"]), 400.0)

        dynamic_name = _matching_parameter(parsed.data.columns, "DYN_LEAK")
        self.assertEqual(parsed.data[dynamic_name].dropna().tolist(), [0.05, 0.075])
        self.assertIn(dynamic_name, parsed.specs["Parameter"].tolist())
        output_text = "|".join(str(column).upper() for column in parsed.data.columns)
        self.assertNotIn("CONT_NEW", output_text)
        self.assertNotIn("CONTACT", output_text)
        self.assertNotIn("SAME", output_text)
        self.assertNotIn("DELAY", output_text)

    def test_old_and_extended_files_merge_as_union_with_missing_values(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            input_dir = root / "input"
            input_dir.mkdir()
            _make_text_source(
                input_dir,
                manufacturing_lot="M000000001-001",
                batch="C000001.00",
            )
            self._long_source(
                input_dir,
                manufacturing_lot="M000000002-001",
                batch="C000002.00",
            )
            cleaner = DianjiDCCleaner(input_dir, root / "output")

            self.assertTrue(cleaner.process_all())
            result = pd.read_excel(cleaner.last_output_file, sheet_name="RAW")

        self.assertEqual(len(result), 6)
        self.assertIn("IDSS80(nA)", result.columns)
        self.assertTrue(
            result.loc[result["批次"] == "C000001.00", "IDSS80(nA)"].isna().all()
        )
        extended = result.loc[result["批次"] == "C000002.00", "IDSS80(nA)"]
        self.assertEqual(extended.dropna().tolist(), [125.0, 250.0])
        self.assertEqual(int(extended.isna().sum()), 1)

    def test_new_known_item_rejects_unsupported_unit_and_missing_bias(self):
        cases = (
            ({"name": "IDSS", "bias1": "VDS=80.00V", "unit": "mV"}, "单位不支持"),
            ({"name": "IDSS", "bias1": "VGS=80.00V", "unit": "uA"}, "缺少 VDS"),
        )
        for item, message in cases:
            with self.subTest(item=item), tempfile.TemporaryDirectory() as temp:
                source = _make_text_source(Path(temp))
                _append_text_items(source, [{**item, "values": {"1": "1"}}])
                with self.assertRaisesRegex(DianjiFormatError, message):
                    parse_powertech_file(source)

    def test_same_display_name_with_different_raw_semantics_fails_closed(self):
        with tempfile.TemporaryDirectory() as temp:
            source = _make_text_source(Path(temp))
            _append_text_items(
                source,
                [
                    {
                        "name": "IDSS",
                        "bias1": "VDS=80.00V",
                        "unit": "nA",
                        "values": {"1": "125"},
                    },
                    {
                        "name": "IDSS80",
                        "unit": "nA",
                        "values": {"1": "125"},
                    },
                ],
            )
            with self.assertRaisesRegex(
                DianjiFormatError, "生成输出列名语义冲突"
            ):
                parse_powertech_file(source)


class PowerTechXlsxDynamicParameterTests(unittest.TestCase):
    def test_includes_right_side_business_item_and_excludes_controls(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as temp:
            source = _make_xlsx_source(Path(temp), layout_name="dj7-39")
            _append_xlsx_items(
                source,
                [
                    {
                        "name": "DYN_CAP",
                        "bias1": "DC Bias=25.00V",
                        "bias3": "Freq=1.00MHz",
                        "unit": "pF",
                        "low": "100",
                        "high": "5000",
                        "values": {"1": 1250.0},
                    },
                    {"name": "CONT_NEW", "unit": "V", "values": {"1": 1.0}},
                    {"name": "SAME", "bias1": "M#=40", "values": {"1": 1.0}},
                    {"name": "DELAY", "bias1": "Time=10ms", "unit": "ms", "values": {"1": 10.0}},
                ],
            )
            parsed = parse_powertech_xlsx_file(source)
            _collect_xlsx_handles()

        dynamic_name = _matching_parameter(parsed.data.columns, "DYN_CAP")
        self.assertEqual(parsed.data[dynamic_name].dropna().tolist(), [1250.0])
        dynamic_spec = parsed.specs.loc[
            parsed.specs["Parameter"] == dynamic_name
        ].iloc[0]
        self.assertEqual(dynamic_spec["Unit"], "pF")
        self.assertEqual(float(dynamic_spec["Low_Limit"]), 100.0)
        self.assertEqual(float(dynamic_spec["High_Limit"]), 5000.0)
        output_text = "|".join(str(column).upper() for column in parsed.data.columns)
        self.assertNotIn("CONT_NEW", output_text)
        self.assertNotIn("SAME", output_text)
        self.assertNotIn("DELAY", output_text)

    def test_required_parameter_unit_and_bias_still_fail_closed(self):
        cases = ((18, 7, "V"), (9, 7, "VCE=16.00V"))
        for row, column, value in cases:
            with self.subTest(row=row, value=value), tempfile.TemporaryDirectory(
                ignore_cleanup_errors=True
            ) as temp:
                source = _make_xlsx_source(Path(temp), layout_name="dj7-39")
                workbook = load_workbook(source)
                workbook["Datalog"].cell(row, column, value)
                workbook.save(source)
                workbook.close()
                message = "单位不支持" if row == 18 else "Bias1 未经验证"
                with self.assertRaisesRegex(DianjiFormatError, message):
                    parse_powertech_xlsx_file(source)
                _collect_xlsx_handles()


class DianjiDynamicRawPatTests(unittest.TestCase):
    def test_raw_pat_unions_parameter_sets_and_counts_available_values(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source_dir = root / "raw"
            source_dir.mkdir()
            _make_text_source(
                source_dir,
                manufacturing_lot="M000000001-001",
                batch="C000001.00",
            )
            long_source = _make_text_source(
                source_dir,
                manufacturing_lot="M000000002-001",
                batch="C000002.00",
            )
            _append_text_items(
                long_source,
                [
                    {
                        "name": "IDSS",
                        "bias1": "VDS=80.00V",
                        "unit": "uA",
                        "values": {"1": "0.125", "4": "0.250"},
                    },
                    {
                        "name": "DYN_PAT",
                        "unit": "V",
                        "values": {"1": "2.5", "4": "3.5"},
                    },
                    {"name": "CONT_NEW", "unit": "V", "values": {"1": "1"}},
                ],
            )

            result = build_raw_pat(
                source_dir,
                spool_dir=root / "spool",
                progress_interval=0,
            )

        stats = result.iloc[1:].set_index("统计量")
        self.assertEqual(int(stats.loc["IDSS80(nA)", "总计数"]), 2)
        dynamic_name = _matching_parameter(stats.index, "DYN_PAT")
        self.assertEqual(int(stats.loc[dynamic_name, "总计数"]), 2)
        self.assertFalse(any("CONT_NEW" in str(name).upper() for name in stats.index))


if __name__ == "__main__":
    unittest.main()
