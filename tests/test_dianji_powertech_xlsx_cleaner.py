import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from factories.dianji.dc_cleaner import DianjiDCCleaner
from factories.dianji.models import DianjiFormatError
from factories.dianji.powertech_xlsx_parser import (
    POWERTECH_XLSX_LAYOUTS,
    is_powertech_xlsx_file,
    parse_powertech_xlsx_file,
    parse_powertech_xlsx_filename,
)
from factories.dianji.source_registry import detect_dianji_source_format


PRODUCT = "NCE40ED120VT(LA)"


def _make_source(
    directory: Path,
    *,
    layout_name: str = "dj7-39",
    manufacturing_lot: str = "M260303020-005",
    batch: str = "FA5Z-9336",
    label: str | None = None,
    trailing_all: bool | None = None,
    changed_item: tuple[int, str] | None = None,
    changed_unit: tuple[int, str] | None = None,
    metadata_batch: str | None = None,
) -> Path:
    layout = next(value for value in POWERTECH_XLSX_LAYOUTS if value.name == layout_name)
    default_variant = next(iter(layout.filename_variants))
    if label is None:
        label = default_variant[0]
    if trailing_all is None:
        trailing_all = default_variant[2]
    tester = default_variant[1]
    label_text = f" {label}" if label else ""
    data_label = "DC" if label == "ALL" else label
    data_label_text = f" {data_label}" if data_label else ""
    lot_label = "QC" if label == "ALL" else label
    path = directory / (
        f"{PRODUCT}_{manufacturing_lot} {batch}{label_text}_FT_"
        f"260320213934_A_{tester}{'ALL' if trailing_all else ''}.xlsx"
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Datalog"
    sheet.cell(1, 1, "PowerTECH Test System")
    sheet.cell(1, 3, "Tester Serial:")
    sheet.cell(1, 4, tester)
    sheet.cell(1, 6, "Station :")
    sheet.cell(1, 7, "A")
    sheet.cell(2, 1, "DataFileName:")
    sheet.cell(
        2,
        3,
        f"D:\\DC数据\\{manufacturing_lot} {batch}{data_label_text}_FT_"
        f"260320213934_A_{tester}.plf",
    )
    sheet.cell(3, 1, "TestFileName:")
    sheet.cell(3, 3, f"Z:\\DC测试程序\\{PRODUCT} ALL M05M06M07M08 需动态.ptf")
    sheet.cell(5, 1, "Lot:")
    sheet.cell(5, 3, f"{manufacturing_lot} {metadata_batch or batch} {lot_label}".strip())

    labels = {
        "Item Name": 8,
        "Bias1": 9,
        "Bias2": 10,
        "Bias3": 11,
        "Min Limit": 12,
        "Max Limit": 13,
        "Serial#": 18,
    }
    for text, row in labels.items():
        sheet.cell(row, 1, text)
    sheet.cell(18, 2, "S#")
    sheet.cell(18, 3, "Bin#")
    for item_number, item_name in enumerate(layout.item_bases, start=1):
        source_name = changed_item[1] if changed_item and changed_item[0] == item_number else item_name
        sheet.cell(8, item_number + 3, f"{item_number} {source_name}")

    for field in layout.output_fields:
        column = field.item_no + 3
        unit = changed_unit[1] if changed_unit and changed_unit[0] == field.item_no else field.unit
        sheet.cell(18, column, unit)
        sheet.cell(12, column, "-1" if field.unit else "-10")
        sheet.cell(13, column, "10")
        for label_name, expected in field.conditions:
            sheet.cell(labels[label_name], column, expected)

    values = {field.item_no: float(index) + 0.25 for index, field in enumerate(layout.output_fields, start=1)}
    for source_row, serial in enumerate((1, 2, 3, 4), start=19):
        sheet.cell(source_row, 1, serial)
        sheet.cell(source_row, 2, 1)
        sheet.cell(source_row, 3, 1)
        for item_number, value in values.items():
            sheet.cell(source_row, item_number + 3, value)
    # Row 2 reached DVCE but failed later; row 3 failed before DVCE; row 4 is overflow.
    for column in range(8, len(layout.item_bases) + 4):
        sheet.cell(20, column).value = None
    sheet.cell(21, 7).value = None
    sheet.cell(22, 7, "over")
    ices_item = next(field.item_no for field in layout.output_fields if field.output_name == "ICES1000(nA)")
    sheet.cell(19, ices_item + 3, 9999)
    workbook.save(path)
    workbook.close()
    return path


class PowerTechXlsxFilenameTests(unittest.TestCase):
    def test_accepts_all_verified_filename_shapes(self):
        names = (
            f"{PRODUCT}_M260108039-001 FA5Y-9413 ALL_FT_260201183535_A_SDTS10212518.xlsx",
            f"{PRODUCT}_M260108039-001-A-B FA5Y-9413_FT_260202132149_A_SDTS10212518ALL.xlsx",
            f"{PRODUCT}_m260303020-005 fa5z-9336 rt_FT_260320213934_A_SDTS10255062.xlsx",
            f"{PRODUCT}_M260303020-005 FA5Z-9336__FT_260319144014_A_SDTS10255062.xlsx",
        )
        for name in names:
            with self.subTest(name=name):
                identity = parse_powertech_xlsx_filename(name)
                self.assertEqual(identity.product, PRODUCT)
                self.assertTrue(identity.batch.startswith("FA"))

    def test_rejects_unknown_product_and_label(self):
        cases = (
            "UNKNOWN_M260303020-005 FA5Z-9336_FT_260320213934_A_SDTS10255062.xlsx",
            f"{PRODUCT}_M260303020-005 FA5Z-9336 M06_FT_260320213934_A_SDTS10255062.xlsx",
        )
        for name in cases:
            with self.subTest(name=name), self.assertRaises(DianjiFormatError):
                parse_powertech_xlsx_filename(name)


class PowerTechXlsxParserTests(unittest.TestCase):
    def test_all_four_layouts_restore_one_output_contract(self):
        expected_columns = [
            "批次", "DVCE(mV)", "Rg(R)", "VTH1(V)", "VTH2(V)",
            "BVDSS1(V)", "BVDSS2(V)", "BVDSS3(V)", "ICES1000(nA)",
            "IGSS30-1(nA)", "ISGS30-1(nA)", "VDSON40A-11V(V)",
            "VDSON40A-15V(V)", "VDSON160A-15V(V)", "VF40A(V)",
            "ICES1200-1(nA)", "ICES1250(nA)", "ICES1200-2(nA)",
            "IGSS30-2(nA)", "ISGS30-2(nA)", "DELTA BV", "DELTA VTH",
        ]
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            for index, layout in enumerate(POWERTECH_XLSX_LAYOUTS):
                source = _make_source(
                    root,
                    layout_name=layout.name,
                    manufacturing_lot=f"M26030302{index}-005",
                    batch=f"FA5Z-{9336 + index}",
                )
                parsed = parse_powertech_xlsx_file(source)
                self.assertEqual(list(parsed.data.columns), expected_columns)
                self.assertEqual(parsed.source_rows, 4)
                self.assertEqual(parsed.kept_rows, 2)
                self.assertTrue(pd.isna(parsed.data.loc[1, "Rg(R)"]))
                self.assertEqual(parsed.invalid_marker_counts["DVCE(mV)"], 1)
                self.assertEqual(parsed.invalid_marker_counts["ICES1000(nA)"], 1)
                self.assertEqual(len(parsed.specs), 21)

    def test_detects_xlsx_signature_and_registry_handler(self):
        with tempfile.TemporaryDirectory() as temp:
            source = _make_source(Path(temp))
            self.assertTrue(is_powertech_xlsx_file(source))
            self.assertEqual(detect_dianji_source_format(source).key, "powertech_xlsx")

    def test_rejects_changed_layout_unit_and_identity(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            cases = (
                (_make_source(root, changed_item=(30, "IDSS")), "Item 布局未经验证"),
                (_make_source(root, manufacturing_lot="M260303021-005", changed_unit=(4, "V")), "单位不支持"),
                (_make_source(root, manufacturing_lot="M260303022-005", metadata_batch="FA5Z-9999"), "Lot 元数据不一致"),
                (
                    _make_source(
                        root,
                        layout_name="dj7-35",
                        manufacturing_lot="M260303023-005",
                        label="DC",
                        trailing_all=False,
                    ),
                    "文件名/机台组合未经验证",
                ),
            )
            for source, message in cases:
                with self.subTest(source=source.name), self.assertRaisesRegex(DianjiFormatError, message):
                    parse_powertech_xlsx_file(source)


class PowerTechXlsxCleanerTests(unittest.TestCase):
    def test_cleaner_writes_raw_and_scatter_bundle(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            input_dir = root / "input"
            input_dir.mkdir()
            _make_source(input_dir, layout_name="dj7-35", batch="FA5Y-9413")
            _make_source(
                input_dir,
                layout_name="dj7-39",
                manufacturing_lot="M260303021-005",
                batch="FA5Z-9336",
            )
            cleaner = DianjiDCCleaner(input_dir, root / "output")
            self.assertTrue(cleaner.process_all())
            result = pd.read_excel(cleaner.last_output_file, sheet_name="RAW")
            self.assertEqual(result.shape, (4, 23))
            self.assertEqual(result["NUM"].tolist(), [1, 2, 3, 4])
            self.assertEqual(cleaner.last_run_summary["source_formats"], {"PowerTECH XLSX": 2})
            self.assertEqual(cleaner.last_run_summary["source_rows"], 8)
            self.assertEqual(cleaner.last_run_summary["kept_rows"], 4)
            self.assertTrue(cleaner.last_scatter_manifest.is_file())


if __name__ == "__main__":
    unittest.main()
